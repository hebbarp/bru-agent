"""
Excel Generator Skill - Creates Excel spreadsheets from data.

Supports:
1. Table data (list of lists or list of dicts)
2. Markdown tables (parsed and converted)
3. Multiple sheets
"""

import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Union
from loguru import logger

from ..base import BaseSkill

# Check for openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger.info(f"Excel generator available - openpyxl: {OPENPYXL_AVAILABLE}")


class ExcelGeneratorSkill(BaseSkill):
    """Skill for generating Excel spreadsheets."""

    name = "create_excel"
    description = """Create an Excel spreadsheet (.xlsx) from data.
Accepts data in these formats:
1. List of lists: [["Header1", "Header2"], ["Row1Col1", "Row1Col2"], ...]
2. List of dicts: [{"Name": "John", "Age": 30}, {"Name": "Jane", "Age": 25}]
3. Markdown table: "| Name | Age |\n|------|-----|\n| John | 30 |"

Use this for tabular data that users want to download and work with in Excel."""
    version = "1.0.0"

    def __init__(self, config: dict = None):
        super().__init__(config)
        self.output_dir = Path(config.get('output_dir', './output')) if config else Path('./output')
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def get_schema(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Title of the Excel document (used in filename and as sheet title)"
                },
                "data": {
                    "oneOf": [
                        {
                            "type": "array",
                            "description": "List of lists or list of dicts representing table data"
                        },
                        {
                            "type": "string",
                            "description": "Markdown table format string"
                        }
                    ],
                    "description": "Table data - can be list of lists, list of dicts, or markdown table string"
                },
                "headers": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Optional column headers (if not included in data)"
                },
                "filename": {
                    "type": "string",
                    "description": "Optional filename for the Excel file (without .xlsx extension)"
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Optional name for the worksheet (default: 'Sheet1')"
                }
            },
            "required": ["title", "data"]
        }

    async def execute(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """Generate an Excel spreadsheet."""
        if not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "openpyxl is not installed. Run: pip install openpyxl"}

        title = params.get('title', 'Untitled Spreadsheet')
        data = params.get('data', [])
        headers = params.get('headers', None)
        filename = params.get('filename', '')
        sheet_name = params.get('sheet_name', 'Sheet1')

        if not data:
            return {"success": False, "error": "No data provided for Excel"}

        # Generate filename if not provided
        if not filename:
            safe_title = "".join(c for c in title if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_title = safe_title.replace(' ', '_')[:50]
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{safe_title}_{timestamp}"

        xlsx_filename = f"{filename}.xlsx"
        filepath = self.output_dir / xlsx_filename

        try:
            # Parse data if it's a markdown table string
            if isinstance(data, str):
                data = self._parse_markdown_table(data)
                if not data:
                    return {"success": False, "error": "Could not parse markdown table"}

            # Convert list of dicts to list of lists with headers
            if data and isinstance(data[0], dict):
                if not headers:
                    headers = list(data[0].keys())
                data = [[row.get(h, '') for h in headers] for row in data]
                data.insert(0, headers)  # Add headers as first row

            # Add provided headers if data doesn't have them
            elif headers and data:
                data.insert(0, headers)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]  # Excel limits sheet names to 31 chars

            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

            # Write data and apply styles
            for row_idx, row_data in enumerate(data, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

                    if row_idx == 1:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    elif row_idx % 2 == 0:  # Alternating row colors
                        cell.fill = alt_fill

            # Auto-adjust column widths
            for col_idx in range(1, len(data[0]) + 1 if data else 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                for row in data:
                    if col_idx <= len(row):
                        cell_value = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ''
                        max_length = max(max_length, len(cell_value))

                adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
                ws.column_dimensions[column_letter].width = max(adjusted_width, 10)

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Save
            wb.save(filepath)

            logger.info(f"Excel file created: {filepath}")

            return {
                "success": True,
                "result": {
                    "message": f"Excel spreadsheet '{title}' created successfully",
                    "filename": xlsx_filename,
                    "filepath": str(filepath),
                    "size_bytes": filepath.stat().st_size,
                    "rows": len(data),
                    "columns": len(data[0]) if data else 0
                }
            }

        except Exception as e:
            logger.error(f"Failed to create Excel: {e}")
            import traceback
            traceback.print_exc()
            return {"success": False, "error": f"Failed to create Excel: {str(e)}"}

    def _parse_markdown_table(self, markdown: str) -> List[List[str]]:
        """Parse a markdown table into a list of lists."""
        lines = markdown.strip().split('\n')
        result = []

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Skip separator rows (|---|---|)
            if re.match(r'^\|[\s\-:|]+\|$', line):
                continue

            # Parse table row
            if line.startswith('|') and line.endswith('|'):
                # Split by | and clean up
                cells = line[1:-1].split('|')
                cells = [cell.strip() for cell in cells]
                result.append(cells)
            elif '|' in line:
                # Handle tables without leading/trailing |
                cells = line.split('|')
                cells = [cell.strip() for cell in cells]
                result.append(cells)

        return result
